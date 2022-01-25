import sys
from openpyxl import load_workbook


# function to generate agent with alternatives from the worksheet 
def generatePreferences(agentsheet):
    agent_vots = []
    dic_vots = {}
    # reading rows of worksheet using for loop
    for row in range(1, agentsheet.max_row + 1):
        agent_vots.append([agentsheet.cell(row, c).value for c in range(1, agentsheet.max_column + 1)])
        #Nested loop reading the each excel sheet cell value and storing in 2d array
    #for loop to read the agent votes and store the results of agent preferences votes in dictionary
    for i in range(len(agent_vots)):
        #get the each index array agent votes from 2d array
        agent_preference = agent_vots[i]
        #sort the agent preferences that will be used to generate the indices
        sorted_vote = []
        uniqe_preference = list(set(agent_preference))
        uniqe_preference.sort(reverse=True)
        for preference in uniqe_preference:
            indeces = [v for v, x in enumerate(agent_preference) if x == preference]
            indeces.sort(reverse=True)
            for index in indeces:
                sorted_vote.append(index + 1)
        agent_vote = sorted_vote
        dic_vots[i + 1] = agent_vote
    return dic_vots


#  A function input preference profile dictionary and agent as integer to return a winner according to Dictatorship rule
def dictatorship(preferenceProfile, agent):
    if agent > len(preferenceProfile) or agent < 1:
        return "Incorrect input"
    else:
        return preferenceProfile[agent][0]


#  A function input preference profile dictionary and tie break to return a winner according to Plurality rule
def plurality(preferenceProfile, tieBreak):
    favorable_candidates = [i[0] for i in preferenceProfile.values()]
    votes_counts = dict((i, favorable_candidates.count(i)) for i in set(favorable_candidates))
    winner_candidates = [t for t in votes_counts if votes_counts[t] == max(votes_counts.values())]
    if len(winner_candidates) > 1:
        winner = tie_break(preferenceProfile, winner_candidates, tieBreak)
    else:
        winner = winner_candidates[0]
    return winner


#  A function input preference profile dictionary and tie break to return a winner according to Veto rule
def veto(preference_profile, tieBreak):
    candiadates = preference_profile[1]
    score_vector = [1] * (len(candiadates) - 1)
    score_vector.append(0)
    the_winner_is = scoringRule(preference_profile, score_vector, tieBreak)
    return the_winner_is


#  A function input preference profile dictionary and tie break to return a winner according to Borda rule
def borda(preference_profile, tieBreak):
    candiadates = preference_profile[1]
    score_vector = list(reversed(range(len(candiadates))))
    the_winner_is = scoringRule(preference_profile, score_vector, tieBreak)
    return the_winner_is


#  A function input preference profile dictionary and tie break to return a winner according to Harmonic rule
def harmonic(preference_profile, tieBreak):
    candiadates = preference_profile[1]
    score_vector = [1/i for i in range(1, len(candiadates) + 1)]
    the_winner_is = scoringRule(preference_profile, score_vector, tieBreak)
    return the_winner_is


#  A function input preference profile dictionary, a score vector, and tie break to return a winner according to score
# and the tie breaker rule
def scoringRule(preference_profile, score_vector, tieBreak):
    if len(preference_profile[1]) == len(score_vector):
        candidate_scores = dict((el, 0) for el in preference_profile[1])
        for i in range(1, len(preference_profile) + 1):
            vote = preference_profile[i]
            for j in range(len(score_vector)):
                candidate = vote[j]
                candidate_scores[candidate] += score_vector[j]
        winner_candidates = [t for t in candidate_scores if candidate_scores[t] == max(candidate_scores.values())]
        if len(winner_candidates) > 1:
            winner = tie_break(preferenceProfile, winner_candidates, tieBreak)
        else:
            winner = winner_candidates[0]
    else:
        print("Incorrect input")
        return False
    return winner


#  A function input preference profile dictionary and tie break to return a winner according to Single Transferable Vote rule
def STV(preference_profile, tieBreak):
    favorable_candidates = []
    votes_counts = ""
    least_fav = []
    while len(preference_profile[1]) > 1:
        no_candidates = len(preference_profile[1])
        favorable_candidates = [i[0] for i in preference_profile.values()]
        votes_counts = [(i, favorable_candidates.count(i)) for i in preference_profile[1]]
        min_votes = min(votes_counts, key=lambda t: t[1])[1]
        least_fav = [t for t in votes_counts if t[1] == min_votes]
        if len(least_fav) != no_candidates:
            for agent in preference_profile.values():
                for least_fav_member in least_fav:
                    agent.remove(least_fav_member[0])
        else:
            break
    winner_candidates = preference_profile[1]
    if len(winner_candidates) > 1:
        winner = tie_break(preferenceProfile, winner_candidates, tieBreak)
    else:
        winner = winner_candidates[0]
    return winner


#  A function input preference profile dictionary and tie break, return a winner according to the maximum sum of
#  numerical values of the xlsx file and in case of tie we use the tie break to distinguish between possible winners
def rangeVoting(ws, tieBreak):
    # create a list of lists with each list will have the values of every row of the sheet
    preference_profile = []
    for row in range(1, ws.max_row + 1):
        preference_profile.append([ws.cell(row, c).value for c in range(1, ws.max_column + 1)])
    vote_values = {i+1: list(preference_profile[i]) for i in range(len(preference_profile))}
    candidates_score = {}
    for i in range(len(vote_values[1])):
        candidate_score = sum([t[i] for t in vote_values.values()])
        candidates_score[i + 1] = candidate_score
    winner_candidates = [t for t in candidates_score if candidates_score[t] == max(candidates_score.values())]
    if len(winner_candidates) > 1:
        winner = tie_break(preferenceProfile, winner_candidates, tieBreak)
    else:
        winner = winner_candidates[0]
    return winner


#  A function input preference profile dictionary, possible winners candidates, and tie break rule, return a winner
#  according to the Tie-breaking rule
def tie_break(preferenceProfile, candiadates, tieBreak):
    if tieBreak == "max":
        the_winner_is = max(candiadates)
    elif tieBreak == "min":
        the_winner_is = min(candiadates)
    else:
        agent_i = preferenceProfile[tieBreak]
        the_winner_is = [t for t in agent_i if t in candiadates][0]
    return the_winner_is
